"""
Export consolidated retrievability analysis reports to Excel spreadsheet.

This script processes consolidated JSON files and creates a formatted Excel workbook
with recommendations grouped by dimension and category.
"""

import os
import sys
import glob
import json
import argparse
from pathlib import Path
from typing import Dict, List, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def load_consolidated_report(file_path: str) -> Optional[Dict]:
    """Load a consolidated JSON report.
    
    Args:
        file_path: Path to consolidated JSON file
        
    Returns:
        Dictionary with report data or None on error
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading {file_path}: {e}")
        return None


def extract_data_from_report(report: Dict) -> List[Dict]:
    """Extract recommendation data from a consolidated report.
    
    Args:
        report: Consolidated report dictionary
        
    Returns:
        List of recommendation dictionaries with file context
    """
    source_file = report.get('source_file', 'Unknown')
    content_type = report.get('content_type', 'Unknown')
    recommendations = report.get('recommendations', [])
    
    if not recommendations:
        # Return a single row indicating no issues
        return [{
            'source_file': source_file,
            'content_type': content_type,
            'category': 'N/A',
            'dimension': 'N/A',
            'evidence': '',
            'action': 'No high-confidence issues found',
            'impact': '',
            'status': '✅ Ready',
        }]
    
    rows = []
    for rec in recommendations:
        rows.append({
            'source_file': source_file,
            'content_type': content_type,
            'category': rec.get('category', 'unknown'),
            'dimension': rec.get('dimension', 'unknown'),
            'evidence': rec.get('evidence', ''),
            'action': rec.get('action', ''),
            'impact': rec.get('impact', ''),
            'status': '',  # For writer to fill in
        })
    
    return rows


def process_consolidated_directory(consolidated_dir: str) -> List[Dict]:
    """Process all consolidated JSON files in a directory.
    
    Args:
        consolidated_dir: Path to consolidated directory
        
    Returns:
        List of extracted data dictionaries
    """
    json_files = glob.glob(os.path.join(consolidated_dir, "*_consolidated_*.json"))
    
    if not json_files:
        print(f"No consolidated JSON files found in {consolidated_dir}")
        return []
    
    print(f"Found {len(json_files)} consolidated reports")
    
    all_rows = []
    for file_path in json_files:
        report = load_consolidated_report(file_path)
        if report:
            rows = extract_data_from_report(report)
            all_rows.extend(rows)
            print(f"  Processed: {os.path.basename(file_path)} ({len(rows)} recommendations)")
    
    return all_rows


def create_excel_with_formatting(df: pd.DataFrame, output_path: str):
    """Create Excel file with formatting.
    
    Args:
        df: DataFrame with all data
        output_path: Output Excel file path
    """
    # Write initial Excel file
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Recommendations', index=False)
    
    # Load workbook for formatting
    wb = load_workbook(output_path)
    ws = wb['Recommendations']
    
    # Define colors for categories
    category_colors = {
        'structural': 'B4C7E7',  # Light blue
        'semantic': 'C5E0B4',     # Light green
        'query': 'FFE699',        # Light yellow
        'redundancy': 'F8CBAD',   # Light orange
        'N/A': 'E2EFDA',          # Light gray-green
    }
    
    # Header formatting
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    # Column widths
    column_widths = {
        'A': 35,  # source_file
        'B': 15,  # content_type
        'C': 15,  # category
        'D': 25,  # dimension
        'E': 60,  # evidence
        'F': 50,  # action
        'G': 40,  # impact
        'H': 15,  # status
    }
    
    # Apply column widths
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Format header row
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Format data rows
    for row_idx in range(2, len(df) + 2):
        category = ws.cell(row=row_idx, column=3).value  # Column C is category
        
        # Apply category color to entire row
        if category in category_colors:
            fill = PatternFill(start_color=category_colors[category], 
                             end_color=category_colors[category], 
                             fill_type='solid')
            for col_idx in range(1, len(df.columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
        
        # Apply text wrapping to evidence, action, impact columns
        for col_idx in [5, 6, 7]:  # E, F, G columns
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Center align category and dimension
        for col_idx in [3, 4]:  # C, D columns
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center', vertical='top')
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    # Save formatted workbook
    wb.save(output_path)
    print(f"\n✓ Excel file created: {output_path}")


def main():
    """Main function to export consolidated reports to Excel."""
    parser = argparse.ArgumentParser(
        description='Export consolidated retrievability analysis reports to Excel spreadsheet'
    )
    parser.add_argument(
        'consolidated_dir',
        nargs='?',
        help='Path to consolidated directory (default: output/test/consolidated)'
    )
    parser.add_argument(
        '--output', '-o',
        help='Output Excel file path (default: auto-generated in output directory)'
    )
    
    args = parser.parse_args()
    
    # Set up paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    if args.consolidated_dir:
        consolidated_dir = args.consolidated_dir
    else:
        consolidated_dir = os.path.join(script_dir, "output/test/consolidated")
    
    if not os.path.exists(consolidated_dir):
        print(f"Error: Consolidated directory not found at {consolidated_dir}")
        return 1
    
    print("=== Retrievability Analysis Excel Exporter ===")
    print(f"Reading from: {consolidated_dir}\n")
    
    # Process all consolidated reports
    all_rows = process_consolidated_directory(consolidated_dir)
    
    if not all_rows:
        print("No data to export")
        return 1
    
    # Create DataFrame
    df = pd.DataFrame(all_rows)
    
    # Reorder columns
    column_order = ['source_file', 'content_type', 'category', 'dimension', 
                    'evidence', 'action', 'impact', 'status']
    df = df[[col for col in column_order if col in df.columns]]
    
    # Determine output path
    if args.output:
        output_path = args.output
    else:
        output_dir = os.path.dirname(consolidated_dir)
        output_path = os.path.join(output_dir, "retrievability_recommendations.xlsx")
    
    # Create Excel file with formatting
    create_excel_with_formatting(df, output_path)
    
    print(f"\n--- Export Complete ---")
    print(f"Total rows: {len(all_rows)}")
    print(f"Files processed: {df['source_file'].nunique()}")
    
    # Summary by category
    category_counts = df[df['category'] != 'N/A']['category'].value_counts()
    if not category_counts.empty:
        print("\nRecommendations by category:")
        for cat, count in category_counts.items():
            print(f"  - {cat}: {count}")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
